# -*- coding: utf-8 -*-
"""proyecciones

Automatically generated by Colaboratory.

Original file is located at
    https://colab.research.google.com/drive/1JfZgkYNrr0rciqog8GCieseLBfTfMuCz
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl import worksheet
#from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl import cell
#openpyxl.utils.cell
#from openpyxl.styles import colors
from openpyxl.styles import PatternFill

# Change to the /content directory
os.chdir('/content')

# Now, you can work with files in the /content directory
# with open('example.txt', 'w') as file:
    # file.write('Hello, Colab!')


#amarilloAños=colors.Color("FFFF00")
#AmarilloPeriodos=colors.Color("FFFFCC")
#verdeReajustes=colors.Color("E5FFCC")


amarilloAños=PatternFill(patternType='solid',fgColor="FFFF00")
amarilloPeriodos=PatternFill(patternType='solid',fgColor="FFFFCC")
verdeReajustes=PatternFill(patternType='solid',fgColor="99FF99")

t="""ej-1	6	2	$B$2	$C$2
ej-2	3	5	$B$3	$C$3
ej-3	4	3	$B$4	$C$4
ej-4	1	7	$B$5	$C$5"""

t=t.split("\n")

for i in range(len(t)):
  t[i]=t[i].split("\t")

periocidadyTasa={}
for i in t:
  periocidadyTasa[i[0]]={"periocidad":int(i[1]),"tasa":int(i[2]),"cperiocidad":i[3],"ctasa":i[4]}

print(periocidadyTasa)

wb = load_workbook(filename = 'prueba.xlsx')
hojas=wb.sheetnames

hojas.remove("centros")

#for row in enumerate(wb["centros"].iter_rows(min_row=2, values_only=True),2):




#periocidad columna 2 y tasa columna 3
#periocidadyTasa=wb["centros"]

#excelpyt=load_workbook(filename= "tasas.xlsx")
#periocidadyTasa=excelpyt["Hoja1"]

#.cell(1,1).value

def sumaCeldas(l):
  b=True
  for i in l:
    if b:
      s="="+i
      b=False
    else:
      s=s+"+"+i
  return s

celdaProyecciones=[]
numeroHoja=0
for hoja in hojas:
  fila=2 #partir en fila 2 por encabezados
  columna=5 #partir en columna 5 por tasas

  #tasas
  wb[hoja].cell(1,4).value="Tasa (%)"
  #wb[hoja].cell(2,4).value=f"={periocidadyTasa.cell(fila+numeroHoja,3).value}/100"
  #t='\"centros\"!'
  t="centros!"
  wb[hoja].cell(2,4).value=f"={t+periocidadyTasa[hoja]['ctasa']}/100"
  celdaTasa=wb[hoja].cell(2,4).coordinate


  #coordenada de la tasa fija
  celdaTasa="$"+celdaTasa[0]+"$"+celdaTasa[1]

  #periocidad=periocidadyTasa.cell(fila+numeroHoja,2).value
  periocidad=periocidadyTasa[hoja]["periocidad"]
  columnaAños=[]

  numeroHoja+=1


  for row in wb[hoja].iter_rows(min_row=2, values_only=True):


    item = row[0]
    precio = row[1]
    cantidad = row[2]
    wb[hoja].cell(1,columna).value="Precio x Cant"

    celdaPrecio=wb[hoja].cell(fila,2).coordinate
    celdaCantidad=wb[hoja].cell(fila,3).coordinate
    celdaCantidad="$"+celdaCantidad[0]+"$"+celdaCantidad[1]

    # precio x cantidad columna n° 4
    wb[hoja].cell(fila,columna).value=f"={celdaPrecio}*{celdaCantidad}"
    columna+=1

    #---recorrer en la planilla (hacia la derecha)
    numReajustesYear=int(12/periocidad)
    nTotalReajustes=numReajustesYear*3

    celdasPeriodos=[]
    contador=0
    year=2024
    celdaReajuste=celdaPrecio

    for i in range(nTotalReajustes):
          wb[hoja].cell(1,columna).value=f"Reajuste-{i+1}"
          wb[hoja].cell(fila,columna).value=f"=({celdaTasa}+1)*{celdaReajuste}"
          celdaReajuste=wb[hoja].cell(fila,columna).coordinate
          columna+=1

          wb[hoja].cell(1,columna).value=f"Periodo {contador+1}"
          wb[hoja].cell(fila,columna).value=f"={celdaReajuste}*{celdaCantidad}*{periocidad}"
          celdasPeriodos.append(wb[hoja].cell(fila,columna).coordinate)
          columna+=1
          contador+=1

          if contador==numReajustesYear:
            wb[hoja].cell(fila,columna).value=sumaCeldas(celdasPeriodos)
            wb[hoja].cell(1,columna).value=f"Total {year}"
            if columna not in columnaAños:
              columnaAños.append(columna)
            wb[hoja].fill=amarilloAños
            columna+=1
            year+=1
            celdasPeriodos=[]
            contador=0
    #CAMBIO DE FILA. ES DECIR, DE Item
    fila+=1
    columna=5
    celdaReajuste=celdaPrecio

  celdaAños=[]
  for c in columnaAños:

    ini=wb[hoja].cell(2,c).coordinate
    fin=wb[hoja].cell(fila-1,c).coordinate
    sSuma=f"=SUMA({ini}:{fin})"

    #wb[hoja].cell(fila,c).value=worksheet.formula.ArrayFormula(wb[hoja].cell(fila,c).coordinate,sSuma)
    wb[hoja].cell(fila,c).value=sSuma
    celdaP=str(hoja)+"!"+wb[hoja].cell(fila,c).coordinate
    #periocidadyTasa
    celdaAños.append(celdaP)



  #celdaProyecciones[str(hoja)]=celdaAños
  celdaProyecciones.append(celdaAños)








wb.save('output_file.xlsx')
print(celdaProyecciones)

for i in celdaProyecciones:
  print("\n")
  for j in i:
    print(f"={j}", end="")
    print(" ",end="")